import re
import json
import torch
from transformers import pipeline, AutoTokenizer, AutoModelForCausalLM
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MODEL_NAME = "google/gemma-3-12b-it"
# MODEL_NAME = "Qwen/Qwen2.5-7B-Instruct"


hf_token = 'token'


OUTPUT_EXCEL = "KG_gemma.xlsx"
SAVE_EVERY_N_ROWS = 3

LAW_ID = "43/2019/QH14"
COL_DOC_ID   = "Doc ID"
COL_DOCUMENT = "Document"
COL_KHOAN    = "Khoản"
COL_DIEU     = "Điều"
COL_NOI_DUNG = "Nội dung"
COL_TAI_LIEU = "Tài liệu"

def load_model(model_name: str = MODEL_NAME):
    print(f"[INFO] Đang tải model: {model_name} ...")
    tokenizer = AutoTokenizer.from_pretrained(
        MODEL_NAME,
        trust_remote_code=True,
        token=hf_token
        # local_files_only=True
    )

    model = AutoModelForCausalLM.from_pretrained(
        MODEL_NAME,
        torch_dtype=torch.bfloat16,
        device_map="auto",
        trust_remote_code=True,
        token=hf_token
        # local_files_only=True
    )
    generator = pipeline(
        "text-generation",
        model=model,
        tokenizer=tokenizer,
    )
    print(f"[INFO] Đã tải model thành công!")
    return generator

def remove_thinking(text: str) -> str:
    """Xóa toàn bộ phần <think>...</think> trong output của Qwen3."""
    return re.sub(r'<think>[\s\S]*?</think>', '', text).strip()


def call_model(generator, prompt: str, max_new_tokens: int = 2048) -> str:
    messages = [
        {
            "role": "system",
            "content": [{"type": "text",
                         "text": "Bạn là một trợ lý AI thông minh. BẮT BUỘC: Toàn bộ câu trả lời phải bằng TIẾNG VIỆT CÓ DẤU đầy đủ. TUYỆT ĐỐI KHÔNG dùng tiếng Anh, tiếng Trung, hay bất kỳ ngôn ngữ nào khác. KHÔNG dùng gạch chân (_), CamelCase, hay viết dính liền các từ. Các từ phải cách nhau bằng dấu cách bình thường."}]
        },
        {
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
            ]
        }
    ]

    outputs = generator(
        messages,
        max_new_tokens=max_new_tokens,
        do_sample=True,
        temperature=0.1,
        top_p=0.95,
        num_return_sequences=1,
    )

    result = outputs[0]["generated_text"]

    if isinstance(result, list):
        return remove_thinking(result[-1]["content"].strip())
    return remove_thinking(result.strip())


def extract_only_questions(text: str) -> str:
    """
    Chỉ giữ lại các dòng bắt đầu bằng CQ kèm theo số (VD: CQ1., CQ2.)
    Loại bỏ toàn bộ các phần râu ria (Tài liệu, ####, Các câu hỏi,...)
    """
    # Tìm tất cả các dòng bắt đầu bằng 'CQ' + số + dấu '.' + nội dung
    # Cờ re.MULTILINE (re.M) giúp ký tự '^' khớp với đầu mỗi dòng
    questions = re.findall(r'^CQ\d+\..*', text, flags=re.MULTILINE)

    # Ghép các câu hỏi lại thành một chuỗi, cách nhau bằng dấu xuống dòng
    return '\n'.join(questions).strip()


def create_CQ(generator, document: str, max_tokens: int = 1024) -> str:
    """
    Bước 1: Tạo các Competency Questions (CQ) từ tài liệu.
    Sử dụng one-shot prompting với ví dụ mẫu.
    """
    prompt = f"""### HƯỚNG DẪN:
⚠️ QUY TẮC BẮT BUỘC:
- Toàn bộ câu hỏi phải bằng **tiếng Việt có dấu đầy đủ** (ví dụ: "khoa học", không phải "khoa hoc" hay "KhoaHoc").
- KHÔNG dùng tiếng Anh, tiếng Trung, hay bất kỳ ngôn ngữ nào khác.
- KHÔNG dùng gạch chân (_) để nối các từ. Các từ phải cách nhau bằng dấu cách.

Hãy viết **các câu hỏi năng lực (competency questions)** dựa trên các **khái niệm trừu tượng cấp cao** trong tài liệu.
Các câu hỏi phải được trả lời DUY NHẤT dựa vào nội dung trong tài liệu.
Viết **tối đa 5 câu hỏi cho mỗi tài liệu (có thể ít hơn 5 câu hỏi)**.

Dưới đây là ví dụ — hãy tuân theo đúng định dạng này khi tạo câu hỏi năng lực,chỉ đưa ra các câu hỏi.

### VÍ DỤ:
####
Tài liệu:
Phạm vi điều chỉnh. Luật này quy định về hệ thống giáo dục quốc dân; cơ sở giáo dục, nhà giáo, người học; quản lý nhà nước về giáo dục; quyền và trách nhiệm của cơ quan, tổ chức, cá nhân liên quan đến hoạt động giáo dục.

####
Các câu hỏi:
CQ1. Luật này quy định về những nội dung nào trong hệ thống giáo dục quốc dân?
CQ2. Luật này điều chỉnh các đối tượng nào trong hoạt động giáo dục?
CQ3. Luật này quy định những nội dung gì liên quan đến quản lý nhà nước về giáo dục?
CQ4. Luật này nêu quyền và trách nhiệm của những chủ thể nào trong hoạt động giáo dục?

---

### NHIỆM VỤ CỦA BẠN:
####
Tài liệu:
{document}
####

Các câu hỏi:"""

    return extract_only_questions(call_model(generator, prompt, max_tokens))


def create_CQ_answer(generator, document: str, cq: str, max_tokens: int = 1024) -> str:
    prompt = f"""### HƯỚNG DẪN:
⚠️ QUY TẮC BẮT BUỘC:
- Toàn bộ câu trả lời phải bằng **tiếng Việt có dấu đầy đủ**.
- KHÔNG dùng tiếng Anh, tiếng Trung, hay bất kỳ ngôn ngữ nào khác.
- KHÔNG dùng gạch chân (_), CamelCase, hay viết dính liền các từ. Các từ cách nhau bằng dấu cách.

Hãy sử dụng tài liệu được cung cấp để trả lời các câu hỏi của người dùng.
Nếu bạn không biết câu trả lời, hãy nói rằng bạn không biết — đừng cố bịa ra câu trả lời.
Trả lời ngắn gọn, chính xác theo nội dung tài liệu.

Tài liệu: {document}

Các câu hỏi: {cq}

### Câu trả lời:"""

    return call_model(generator, prompt, max_tokens)

ontology = """
1. Các lớp 
LEGAL_PROVISION: Đơn vị cấu trúc của văn bản pháp luật, bao gồm:
•	Chương
•	Mục
•	Tiểu mục
•	Điều
•	Khoản
•	Điểm
ORG: Tổ chức cơ quan nhà nước, cơ sở giáo dục, … 
PERSON: Cá nhân, bao gồm: 
•	Người học 
•	Người dạy 
•	Vai trò, chức danh quản lý 
•	Nhóm còn lại: Công dân, người dân Việt Nam. 
EDU_SYSTEM: Các thành phần tạo nên bộ khung của hệ thống giáo dục: 
•	Hệ thống/ cấp học/ trình độ học: Hệ thống giáo dục quốc dân, giáo dục chính quy, giáo dục mầm non, giáo dục đại học, trình độ thạc sĩ. 
•	Chương trình, đơn vị học tập: Chương trình giáo dục phổ thông, môn học, tín chỉ, khối lượng kiến thức. 
•	Khối lớp: Lớp một, lớp mười, mầm non. 
RULE: Các khái niệm trừu tượng mang tính chất định hướng, ràng buộc, cấm đoán, bao gồm: 
•	Phẩm chất / Tiêu chuẩn: Đạo đức, tri thức, sức khỏe, chuẩn nghề nghiệp, chuyên môn nghiệp vụ. 
•	Quyền / Nghĩa vụ / Trách nhiệm: Quyền tự chủ, trách nhiệm giải trình, tôn trọng nhân phẩm. 
•	Hành vi (Được phép hoặc Bị cấm): Gian lận thi cử, xuyên tạc nội dung giáo dục, ép buộc học sinh học thêm. 
•	Mục tiêu vĩ mô: Nâng cao dân trí, bồi dưỡng nhân tài. 
ASSET: Các yếu tố, tài sãn hữu hình có giá trị trao đổi, cấp phát, đầu tư bao gồm: 
•	Văn bằng, chứng chỉ: Bằng tốt nghiệp, … 
•	Tài chính: Học phí, học bổng, ngân sách nhà nước, … 
•	Cơ sở vật chất: sách giáo khoa, đất đai, thiết bị dạy học. 
METRIC: Chỉ số đo lường dùng làm mốc cho các điều kiện pháp lý, bao gồm: 
•	Độ tuổi 
•	Thời gian, niên hạn 
•	Định mức, định lượng 
2. Relation Types 
BAO GỒM: Quan hệ phân cấp. (VD: giáo dục phổ thông bao gồm tiểu học, THCS, THPT) 
•	Domain: EDU_SYSTEM / ORG / LEGAL_PROVISION # Lớp nguồn 
•	Range: EDU_SYSTEM / ORG / LEGAL_PROVISION # Lớp đích 
QUY ĐỊNH VỀ: Liên kết một đơn vị cấu trúc của văn bản pháp luật với nội dung, thực thể hoặc quy tắc mà nó xác lập hoặc điều chỉnh.
•	Domain: LEGAL_PROVISION # Lớp nguồn
•	Range: ORG / PERSON / EDU_SYSTEM / RULE / ASSET / METRIC # Lớp đích
ĐƯỢC ĐỊNH NGHĨA LÀ: Giải thích một từ khóa, thuật ngữ 
•	Domain: ORG / PERSON / EDU_SYSTEM / RULE / ASSET / METRIC # Lớp nguồn 
•	Range: ORG / PERSON / EDU_SYSTEM / RULE / ASSET / METRIC # Lớp đích 
ĐIỀU HÀNH BỞI: Quyền hạn quản lý nhà nước 
•	Domain: EDU_SYSTEM / ORG # Lớp nguồn 
•	Range: ORG # Lớp đích 
CÓ ĐIỀU KIỆN: Nghĩa vụ, điều kiện, tiêu chuẩn. 
•	Domain: EDU_SYSTEM / ORG / PERSON / ASSET # Lớp nguồn 
•	Range: RULE / METRIC # Lớp đích 
CÓ QUYỀN: Quyền lợi hợp pháp 
•	Domain: PERSON / ORG # Lớp nguồn 
•	Range: RULE / ASSET # Lớp đích 
NGHIÊM CẤM: Ngiêm cấm 
•	Domain: PERSON / ORG # Lớp nguồn 
•	Range: RULE # Lớp đích 
BAN HÀNH: Ban hành 
•	Domain: ORG # Lớp nguồn 
•	Range: ASSET / RULE / LEGAL_PROVISION # Lớp đích 
ÁP DỤNG CHO: Áp dụng cho 
•	Domain: RULE / ASSET # Lớp nguồn 
•	Range: PERSON / ORG / EDU_SYSTEM # Lớp đích 
HỌC VÀ DẠY TẠI: Học và dạy tại 
•	Domain: PERSON # Lớp nguồn 
•	Range: ORG # Lớp đích
"""


def extract_json_array(text: str) -> str:
    """
    Tìm và trích xuất phần văn bản nằm giữa dấu ngoặc vuông [ ] (bao gồm cả ngoặc).
    Loại bỏ các ký tự Markdown thừa như ```json hay ```.
    """
    # Pattern \[[\s\S]*\] sẽ tìm đoạn bắt đầu bằng [, kết thúc bằng ],
    # và bao gồm mọi ký tự ở giữa (kể cả dấu xuống dòng)
    match = re.search(r'\[[\s\S]*\]', text)

    if match:
        return match.group(0).strip()

    # Nếu không tìm thấy ngoặc vuông (trường hợp model trả về lỗi), trả về chuỗi gốc
    return text.strip()

def create_qa_pair(question: str, answer: str) -> str:
    """Ghép câu hỏi và câu trả lời thành cặp Q&A."""
    question_list = [q for q in question.strip().split('\n') if q.strip()]
    answer_list = [a for a in answer.strip().split('\n') if a.strip()]

    qa_pairs = []
    for i in range(min(len(question_list), len(answer_list))):
        qu = re.sub(r'^CQ\d+\.\s*', '', question_list[i].strip())
        an = re.sub(r'^CQ\d+\.\s*', '', answer_list[i].strip())
        if qu and an:
            qa_pairs.append(f"Question: {qu}\n Answer: {an}")

    return "\n".join(qa_pairs)


def create_KG(generator, document: str, question: str, answer: str,
              ontology: str, max_tokens: int = 4096) -> str:

    qa_pairs = create_qa_pair(question, answer)

    prompt = f"""### HƯỚNG DẪN:
Bạn là một **công cụ tạo đồ thị tri thức (Knowledge Graph)** có độ thông minh cao.
Nhiệm vụ của bạn là **tạo ra đồ thị tri thức theo định dạng JSON** dựa trên `Tài liệu nguồn` và `Các câu hỏi và câu trả lời từ tài liệu` được cung cấp.

Bạn **PHẢI tuyệt đối tuân thủ** các quy tắc sau:

1. **Sử dụng Ontology**: Thuộc tính `predicate`(quan hệ) của mỗi bộ ba PHẢI là một trong các `Relation Types` của thuộc tính được định nghĩa trong `Ontology`, **NGHIÊM CẤM BỊA THÊM**.
2. **Tuân theo Lược đồ**: Chủ thể và đối tượng phải tương ứng với `domain` và `range` trong ontology, trong trường hợp không được chỉ định trong ontology thì phải để là 'UNKNOWN'.
3. **Dựa trên sự thật**: Tất cả thông tin PHẢI xuất phát từ duy nhất nội dung Tài liệu nguồn.
4. **Tập trung vào Q&A**: Ưu tiên trích xuất thông tin từ các cặp Hỏi – Đáp đã cung cấp.

---

### VÍ DỤ:
####
Tài liệu:
Luật này quy định về hệ thống giáo dục quốc dân; cơ sở giáo dục, nhà giáo, người học; quản lý nhà nước về giáo dục; quyền và trách nhiệm của cơ quan, tổ chức, cá nhân liên quan đến hoạt động giáo dục.

####
Các cặp câu hỏi, câu trả lời:
Question: Luật này quy định về những nội dung nào trong hệ thống giáo dục quốc dân?
Answer: Luật này quy định về hệ thống giáo dục quốc dân, cơ sở giáo dục, nhà giáo, người học; quản lý nhà nước về giáo dục.
Question: Luật này nêu quyền và trách nhiệm của những chủ thể nào trong hoạt động giáo dục?
Answer: Luật này nêu quyền và trách nhiệm của cơ quan, tổ chức, cá nhân liên quan đến học động giáo dục.

####
Ontology:
{ontology}

####
KG Json:
[
  {{
    "subject": "Luật này",
    "subject_type": "RULE",
    "predicate": "ÁP DỤNG CHO",
    "object": "Hệ thống giáo dục quốc dân",
    "object_type": "EDU_SYSTEM"
  }},
  {{
    "subject": "Luật này",
    "subject_type": "RULE",
    "predicate": "ÁP DỤNG CHO",
    "object": "Cơ sở giáo dục",
    "object_type": "ORG"
  }},
  {{
    "subject": "Luật này",
    "subject_type": "RULE",
    "predicate": "ÁP DỤNG CHO",
    "object": "Nhà giáo",
    "object_type": "PERSON"
  }},
  {{
    "subject": "Luật này",
    "subject_type": "RULE",
    "predicate": "ÁP DỤNG CHO",
    "object": "Người học",
    "object_type": "PERSON"
  }},
  {{
    "subject": "Luật này",
    "subject_type": "RULE",
    "predicate": "ÁP DỤNG CHO",
    "object": "Quản lý nhà nước về giáo dục",
    "object_type": "ORG"
  }},
  {{
    "subject": "Luật này",
    "subject_type": "RULE",
    "predicate": "ÁP DỤNG CHO",
    "object": "Quyền và trách nhiệm",
    "object_type": "UNKNOWN"
  }},
  {{
    "subject": "Quyền và trách nhiệm",
    "subject_type": "RULE",
    "predicate": "ÁP DỤNG CHO",
    "object": "Cơ quan",
    "object_type": "ORG"
  }},
  {{
    "subject": "Quyền và trách nhiệm",
    "subject_type": "RULE",
    "predicate": "ÁP DỤNG CHO",
    "object": "Tổ chức",
    "object_type": "ORG"
  }},
  {{
    "subject": "Quyền và trách nhiệm",
    "subject_type": "RULE",
    "predicate": "ÁP DỤNG CHO",
    "object": "Cá nhân",
    "object_type": "PERSON"
  }}
]

### NHIỆM VỤ CỦA BẠN:
####
Tài liệu: {document}
####
Các cặp Hỏi – Đáp:
{qa_pairs}
####
Ontology JSON: {ontology}
####

KG Triplets:"""

    return extract_json_array(call_model(generator, prompt, max_tokens))


def build_node_id(khoan, dieu) -> str:
    parts = []
    if pd.notna(khoan) and str(khoan).strip() not in ("", "nan"):
        val_khoan = int(khoan) if isinstance(khoan, (int, float)) else str(khoan).strip()
        parts.append(f"khoản{val_khoan}")
    if pd.notna(dieu) and str(dieu).strip() not in ("", "nan"):
        val_dieu = int(dieu) if isinstance(dieu, (int, float)) else str(dieu).strip()
        parts.append(f"điều{val_dieu}")
    parts.append(LAW_ID)
    return "-".join(parts)


def extract_first_subject_from_kg(kg_json_str: str, fallback_document: str) -> str:
    """
    Trích xuất subject đầu tiên từ KG JSON do LM sinh ra.
    Nếu parse JSON thất bại, dùng cụm từ đầu tiên của document làm fallback.
    """
    # Thử parse JSON từ output của LM
    try:
        match = re.search(r'\[[\s\S]*\]', kg_json_str)
        if match:
            triplets = json.loads(match.group(0))
            if triplets and isinstance(triplets, list):
                first = triplets[0]
                subject = first.get("subject", "").strip()
                if subject:
                    return subject
    except (json.JSONDecodeError, KeyError, IndexError):
        pass

    # Fallback: lấy cụm từ đến dấu phẩy / chấm phẩy / dấu chấm đầu tiên
    fallback = re.split(r'[,;.]', fallback_document.strip())[0].strip()
    # Giới hạn độ dài hợp lý
    return fallback[:80] if fallback else "Chủ thể không xác định"


def extract_first_subject_type_from_kg(kg_json_str: str, fallback: str = "UNKNOWN") -> str:
    """
    Trích xuất subject_type của subject đầu tiên từ KG JSON do LM sinh ra.
    Nếu parse JSON thất bại, trả về fallback mặc định là "UNKNOWN".
    """
    try:
        match = re.search(r'\[[\s\S]*\]', kg_json_str)
        if match:
            triplets = json.loads(match.group(0))
            if triplets and isinstance(triplets, list):
                first = triplets[0]
                subject_type = first.get("subject_type", "").strip()
                if subject_type:
                    return subject_type
    except (json.JSONDecodeError, KeyError, IndexError):
        pass

    return fallback


def create_fixed_triplets(khoan, dieu, noi_dung: str, first_subject: str, first_subject_type:str) -> list:
    """
    Tạo 3 triplet cố định để gắn kết toàn bộ bộ luật:

    1. (luật 43/2019/QH14, BAO GỒM, khoản{k}-điều{d}-43/2019/QH14)
    2. (khoản{k}-điều{d}-43/2019/QH14, CÓ NỘI DUNG, {nội dung})
    3. (khoản{k}-điều{d}-43/2019/QH14, QUY ĐỊNH VỀ, {subject đầu tiên})
    """
    node_id = build_node_id(khoan, dieu)

    fixed = [
        {
            "subject": f"luật {LAW_ID}",
            "subject_type": "LEGAL_PROVISION",
            "predicate": "BAO GỒM",
            "object": node_id,
            "object_type": "LEGAL_PROVISION"
        },
        {
            "subject": node_id,
            "subject_type": "LEGAL_PROVISION",
            "predicate": "CÓ NỘI DUNG",
            "object": str(noi_dung).strip(),
            "object_type": "UNKNOWN"
        },
        {
            "subject": node_id,
            "subject_type": "LEGAL_PROVISION",
            "predicate": "QUY ĐỊNH VỀ",
            "object": first_subject,
            "object_type": first_subject_type
        }
    ]
    return fixed


def KG_triplet_pipeline(generator, document: str,
                        khoan=None, dieu=None, noi_dung: str = "") -> dict:
    separator = "=" * 70

    # --- Tạo CQ ---
    # print(f"\n{separator}")
    # print("[BƯỚC 1] Tạo Competency Questions (CQ)...")
    # print(separator)
    cq = create_CQ(generator, document)
    # print(cq)

    # --- Trả lời CQ ---
    # print(f"\n{separator}")
    # print("[BƯỚC 2] Trả lời CQ...")
    # print(separator)
    answer = create_CQ_answer(generator, document, cq)
    # print(answer)

    # --- Tạo KG Triplets bằng LM ---
    # print(f"\n{separator}")
    # print("[BƯỚC 3] Tạo KG Triplets (LM)...")
    # print(separator)
    triplets_text = create_KG(generator, document, cq, answer, ontology)
    # print("\n📊 KG Triplets (LM raw):")
    # print(triplets_text)


    # --- Parse LM triplets thành list dict ---
    lm_triplets = []
    try:
        match = re.search(r'\[[\s\S]*\]', triplets_text)
        if match:
            lm_triplets = json.loads(match.group(0))
    except (json.JSONDecodeError, AttributeError):
        print("[WARN] Không parse được JSON từ LM output, lưu dạng raw string.")
    # print(f"\n{separator}")
    # print(lm_triplets)

    # --- Trích xuất subject đầu tiên từ KG của LM ---
    first_subject = extract_first_subject_from_kg(triplets_text, document)
    # print(f"\n🔑 Subject đầu tiên (dùng cho QUY ĐỊNH VỀ): {first_subject}")
    first_subject_type = extract_first_subject_type_from_kg(triplets_text)

    # --- Tạo 3 triplet cố định ---
    # print(f"\n{separator}")
    # print("[BƯỚC 4] Tạo Triplets CỐ ĐỊNH gắn kết bộ luật...")
    # print(separator)
    fixed_triplets = create_fixed_triplets(khoan, dieu, noi_dung, first_subject, first_subject_type)
    # for t in fixed_triplets:
    #     print(f"  ✅ ({t['subject']}, {t['predicate']}, {t['object']})")

    # --- Gộp tất cả triplets ---
    all_triplets = fixed_triplets + (lm_triplets if isinstance(lm_triplets, list) else [])

    return {
        "competency_questions": cq,
        "cq_answers": answer,
        "triplets_lm_raw": triplets_text,
        "triplets_fixed": fixed_triplets,
        "triplets_all": all_triplets
    }

def export_excel(df: pd.DataFrame, all_results: list):
    triplet_col = []
    json_col    = []

    for res in all_results:
        lines = []
        for t in res.get("triplets_all", []):
            s = t.get("subject", "")
            p = t.get("predicate", "")
            o = t.get("object", "")
            lines.append(f"({s}, {p}, {o})")
        triplet_col.append("\n".join(lines))
        json_col.append(res.get("triplets_all", ""))

    df_out = df.copy()
    df_out["Triplet"]   = triplet_col
    df_out["Json gốc"] = json_col

    df_out.to_excel(OUTPUT_EXCEL, index=False, engine="openpyxl")

    print(f"✅ Đã xuất {len(df_out)} dòng ra: {OUTPUT_EXCEL}")

def main():
    generator = load_model(MODEL_NAME)

    df = pd.read_excel("Data_chunked.xlsx")

    all_results = []

    for idx, row in df.iterrows():
        doc_id = row.get(COL_DOC_ID, idx + 1)
        document = str(row[COL_DOCUMENT]).strip()
        khoan = row.get(COL_KHOAN)
        dieu = row.get(COL_DIEU)
        noi_dung = str(row[COL_NOI_DUNG]).strip() if COL_NOI_DUNG in row and pd.notna(row[COL_NOI_DUNG]) else document

        print("\n" + "🔷" * 35)
        print(f"  📄 DOC {doc_id}  |  Khoản: {khoan}  |  Điều: {dieu}")
        print("🔷" * 35)
        print(document)

        result = KG_triplet_pipeline(
            generator,
            document=document,
            khoan=khoan,
            dieu=dieu,
            noi_dung=noi_dung
        )

        all_results.append({
            "doc_id": doc_id,
            "document": document,
            "khoan": khoan,
            "dieu": dieu,
            **result
        })

        if (idx + 1) % SAVE_EVERY_N_ROWS == 0:
            print(f"\n[HỆ THỐNG] Đang lưu tạm thời kết quả tại dòng {idx + 1}...")
            # Lấy phần DataFrame tương ứng với số kết quả hiện có
            df_subset = df.iloc[:len(all_results)]
            export_excel(df_subset, all_results)

    print("\n✅ Hoàn thành toàn bộ pipeline!")
    export_excel(df, all_results)


if __name__ == "__main__":
    main()