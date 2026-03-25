import re
import json
import torch
from transformers import pipeline, AutoTokenizer, AutoModelForCausalLM
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MODEL_NAME = "google/gemma-3-12b-it"

hf_token = 'token'

local_model_path = "/data2/cmdir/home/ioit107/.cache/huggingface/hub/models--google--gemma-3-12b-it/snapshots/96b6f1eccf38110c56df3a15bffe176da04bfd80"


OUTPUT_EXCEL = "KG_gemma2_no_cq.xlsx"
SAVE_EVERY_N_ROWS = 3

LAW_ID = "43/2019/QH14"
COL_DOC_ID   = "Doc ID"
COL_DOCUMENT = "Document"
COL_KHOAN    = "Khoản"
COL_DIEU     = "Điều"
COL_NOI_DUNG = "Nội dung"
COL_TAI_LIEU = "Tài liệu"

def load_model(model_name: str = MODEL_NAME):
    print(f"[INFO] Đang tải model: {model_name} ...")
    tokenizer = AutoTokenizer.from_pretrained(
        local_model_path,
        trust_remote_code=True,
        # token=hf_token,
        local_files_only=True
    )

    model = AutoModelForCausalLM.from_pretrained(
        local_model_path,
        torch_dtype=torch.bfloat16,
        device_map="auto",
        trust_remote_code=True,
        # token=hf_token,
        local_files_only=True
    )

    generator = pipeline(
        "text-generation",
        model=model,
        tokenizer=tokenizer,
    )
    print(f"[INFO] Đã tải model thành công!")
    return generator

def remove_thinking(text: str) -> str:
    """Xóa toàn bộ phần <think>...</think> trong output."""
    return re.sub(r'<think>[\s\S]*?</think>', '', text).strip()


def call_model(generator, prompt: str, max_new_tokens: int = 2048) -> str:
    messages = [
        {
            "role": "system",
            "content": [{"type": "text",
                         "text": "Bạn là một trợ lý AI thông minh. BẮT BUỘC: Toàn bộ câu trả lời phải bằng TIẾNG VIỆT CÓ DẤU đầy đủ. TUYỆT ĐỐI KHÔNG dùng tiếng Anh, tiếng Trung, hay bất kỳ ngôn ngữ nào khác."}]
        },
        {
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
            ]
        }
    ]

    outputs = generator(
        messages,
        max_new_tokens=max_new_tokens,
        do_sample=True,
        temperature=0.1,
        top_p=0.95,
        num_return_sequences=1,
    )

    result = outputs[0]["generated_text"]

    if isinstance(result, list):
        return remove_thinking(result[-1]["content"].strip())
    return remove_thinking(result.strip())


ontology = """
I. Entity Classes
1. Nhóm Cấu trúc Văn bản (Document Structure)
VanBanLuat: Dùng để định danh tên gọi chính thức của các đạo luật, bộ luật, hiến pháp hoặc các văn bản quy phạm pháp luật khác được dẫn chiếu trong tài liệu. (VD: Luật Giáo dục 2019, Hiến pháp nước Cộng hòa xã hội chủ nghĩa Việt Nam, Luật Trẻ em)
CauTrucChuong: Nhận diện các đơn vị phân chia bố cục cấp cao nhất của văn bản luật. (VD: Chương II, Chương III)
CauTrucDieu: Đơn vị cơ sở chứa đựng quy phạm pháp luật. (VD: Điều 4, Điều 12)
CauTrucKhoan: Đơn vị cơ sở chứa đựng quy phạm pháp luật nằm trong Điều. (VD: Khoản 4, Khoản 12)
2. Nhóm Tổ chức & Thiết chế (Organizations & Institutions)
CoQuanQuanLy: Các chủ thể nắm quyền lập pháp, hành pháp hoặc quyền quản lý chuyên ngành nhà nước. (VD: Quốc hội, Chính phủ, Bộ Giáo dục và Đào tạo, Ủy ban nhân dân cấp tỉnh)
CoSoGiaoDuc: Các tổ chức/đơn vị trực tiếp thực hiện hoạt động giảng dạy, đào tạo. (VD: Trường tiểu học, Trường trung học phổ thông, trung tâm giáo dục thường xuyên)
LoaiHinhTruong: Phân loại cơ sở giáo dục dựa trên nguồn gốc vốn đầu tư, sở hữu và cơ chế vận hành tài chính. (VD: Trường công lập, Trường dân lập, Trường tư thục)
HoiDongToChuc: Các ban, nhóm, hội đồng, hoặc tổ chức được thành lập hợp pháp. (VD: Hội đồng trường, Hội đồng quốc gia thẩm định sách giáo khoa)
3. Nhóm Con người (Personnel/Actors)
NguoiHoc: Chủ thể trung tâm tiếp nhận dịch vụ giáo dục. (VD: Trẻ em, Học sinh, Sinh viên, Nghiên cứu sinh)
NhaGiao: Những người làm nhiệm vụ giảng dạy, giáo dục trực tiếp tại cơ sở giáo dục. (VD: Giáo viên, Giảng viên)
CanBoQuanLy: Cá nhân đảm nhiệm chức vụ điều hành, quản trị nội bộ tại cơ sở giáo dục. (VD: Cán bộ quản lý giáo dục, Hiệu trưởng)
LanhDaoNhaNuoc: Các chức danh đại diện đứng đầu bộ máy nhà nước. (VD: Thủ tướng Chính phủ, Bộ trưởng Bộ Giáo dục và Đào tạo)
4. Nhóm Chuyên môn Giáo dục (Educational Concepts)
HeThongGiaoDuc: Các nhánh lớn phân chia phương thức tổ chức hệ thống giáo dục quốc gia. (VD: Giáo dục chính quy, Giáo dục thường xuyên)
CapHocTrinhDo: Tên gọi các bậc thang kiến thức được phân định theo độ tuổi hoặc cấp độ chuyên môn. (VD: Giáo dục mầm non, Giáo dục tiểu học, Thạc sĩ, Tiến sĩ)
ChuongTrinhHocLieu: Các tài liệu quy định nội dung, cấu trúc giảng dạy. (VD: Chương trình giáo dục phổ thông, sách giáo khoa, giáo trình)
VanBangChungChi: Các chứng nhận có giá trị pháp lý minh chứng việc hoàn thành khóa học. (VD: Bằng tốt nghiệp THPT, Bằng cử nhân)
DonViDoLuong: Các đại lượng quy chuẩn dùng để tính toán thời lượng, khối lượng học tập. (VD: Năm học, Niên chế, Tín chỉ, Mô-đun)
HanhViBiNghiemCam: Các hành vi lệch chuẩn mà pháp luật tuyệt đối cấm. (VD: Gian lận trong thi cử, Xúc phạm danh dự nhà giáo)
YeuCauTinhChat: Các tiêu chuẩn, điều kiện, mục tiêu mà thực thể/hoạt động giáo dục phải đạt được. (VD: tính khoa học thực tiễn, kế thừa, liên thông)
QuyenLoiPhapLy: Đặc quyền, lợi ích, quyền tự chủ mà pháp luật bảo hộ. (VD: Được giáo dục, học tập, học vượt lớp, hưởng chính sách ưu tiên)
CoSoVatChat: Tài sản vật lý, hạ tầng hoặc trang thiết bị phục vụ giáo dục. (VD: cơ sở vật chất, thư viện, trang thiết bị)

II. Relation Types
bao_gom: VanBanLuat -> CauTrucChuong / CauTrucDieu / CauTrucKhoan, CauTrucChuong -> CauTrucDieu / CauTrucKhoan, CauTrucDieu -> CauTrucKhoan, [Bất kỳ Class nào] -> [Class tương ứng]
quy_dinh: VanBanLuat / CauTrucDieu / CauTrucKhoan -> [Bất kỳ Class nào]
ban_hanh: LanhDaoNhaNuoc / CoQuanQuanLy -> ChuongTrinhHocLieu / VanBanLuat
quan_ly_nha_nuoc: LanhDaoNhaNuoc / CoQuanQuanLy -> HeThongGiaoDuc / CoSoGiaoDuc
co_loai_hinh: CoSoGiaoDuc -> LoaiHinhTruong
co_hoi_dong: CoSoGiaoDuc -> HoiDongToChuc
to_chuc_thuc_hien: CoSoGiaoDuc -> ChuongTrinhHocLieu
hoc_tap_tai: NguoiHoc -> CoSoGiaoDuc
giang_day_tai: NhaGiao -> CoSoGiaoDuc
dieu_hanh: CanBoQuanLy -> CoSoGiaoDuc
bao_gom_cap_hoc: HeThongGiaoDuc -> CapHocTrinhDo
su_dung_chuong_trinh: CapHocTrinhDo -> ChuongTrinhHocLieu
do_luong_bang: ChuongTrinhHocLieu / CapHocTrinhDo -> DonViDoLuong
cap_van_bang: CapHocTrinhDo -> VanBangChungChi
duoc_cap: NguoiHoc -> VanBangChungChi
phai_dap_ung: [Bất kỳ Thực thể nào] -> YeuCauTinhChat
nghiem_cam: [Nhóm con người / Tổ chức] -> HanhViBiNghiemCam
co_nhiem_vu: [Nhóm Con người / Tổ chức] -> YeuCauTinhChat / ChuongTrinhHocLieu
duoc_huong: [Nhóm Con người / Tổ chức] -> QuyenLoiPhapLy
"""


def extract_json_array(text: str) -> str:
    """
    Tìm và trích xuất phần văn bản nằm giữa dấu ngoặc vuông [ ] (bao gồm cả ngoặc).
    Loại bỏ các ký tự Markdown thừa như ```json hay ```.
    """
    match = re.search(r'\[[\s\S]*\]', text)

    if match:
        return match.group(0).strip()

    return text.strip()

def create_KG(generator, document: str,
              ontology: str, max_tokens: int = 4096) -> str:

    prompt = f"""### HƯỚNG DẪN:
Bạn là một **công cụ tạo đồ thị tri thức (Knowledge Graph)** có độ thông minh cao.
Nhiệm vụ của bạn là **tạo ra đồ thị tri thức theo định dạng JSON** dựa trên `Tài liệu nguồn` được cung cấp.

Bạn **PHẢI tuyệt đối tuân thủ** các quy tắc sau:

1. **Tuân thủ Ontology tuyệt đối**: Thuộc tính `predicate` (quan hệ) của mỗi bộ ba PHẢI là một trong các `Relation Types` được định nghĩa trong `Ontology`. Thuộc tính `subject_type` và `object_type` PHẢI là một trong các `Entity Classes` trong `Ontology`. **NGHIÊM CẤM BỊA THÊM** relation hoặc entity type mới. Nếu một thực thể không thuộc bất kỳ class nào trong Ontology thì **BỎ QUA**, không tạo triplet cho nó.
2. **Tuân theo Lược đồ**: Chủ thể và đối tượng phải tương ứng với `domain` và `range` trong ontology.
3. **Entity ngắn gọn**: Giá trị `subject` và `object` phải là **cụm danh từ ngắn gọn, súc tích** (tối đa 5-7 từ). KHÔNG copy nguyên câu dài từ tài liệu. Trích xuất khái niệm cốt lõi.
4. **Dựa trên sự thật**: Tất cả thông tin PHẢI xuất phát từ duy nhất nội dung Tài liệu nguồn.

---

### VÍ DỤ:
####
Tài liệu:
Điều 31: Chương trình giáo dục phổ thông. 1. Chương trình giáo dục phổ thông phải bảo đảm các yêu cầu sau đây: Thể hiện mục tiêu giáo dục phổ thông; Quy định yêu cầu về phẩm chất và năng lực của học sinh cần đạt được sau mỗi cấp học, nội dung giáo dục bắt buộc đối với tất cả học sinh trong cả nước; Quy định phương pháp, hình thức tổ chức hoạt động giáo dục và đánh giá kết quả giáo dục đối với các môn học ở mỗi lớp, mỗi cấp học của giáo dục phổ thông; Thống nhất trong cả nước và được tổ chức thực hiện linh hoạt, phù hợp với điều kiện cụ thể của địa phương và cơ sở giáo dục phổ thông; Được lấy ý kiến rộng rãi các tổ chức, cá nhân và thực nghiệm trước khi ban hành; được công bố công khai sau khi ban hành.
####
Ontology: {ontology}
####
KG Json:
[
  {{"subject": "Điều 31", "subject_type": "CauTrucDieu", "predicate": "quy_dinh", "object": "Chương trình giáo dục phổ thông", "object_type": "ChuongTrinhHocLieu"}},
  {{"subject": "Chương trình giáo dục phổ thông", "subject_type": "ChuongTrinhHocLieu", "predicate": "phai_dap_ung", "object": "thể hiện mục tiêu giáo dục phổ thông", "object_type": "YeuCauTinhChat"}},
  {{"subject": "Chương trình giáo dục phổ thông", "subject_type": "ChuongTrinhHocLieu", "predicate": "phai_dap_ung", "object": "yêu cầu về phẩm chất và năng lực", "object_type": "YeuCauTinhChat"}},
  {{"subject": "Chương trình giáo dục phổ thông", "subject_type": "ChuongTrinhHocLieu", "predicate": "phai_dap_ung", "object": "nội dung giáo dục bắt buộc", "object_type": "YeuCauTinhChat"}},
  {{"subject": "Chương trình giáo dục phổ thông", "subject_type": "ChuongTrinhHocLieu", "predicate": "phai_dap_ung", "object": "thống nhất trong cả nước", "object_type": "YeuCauTinhChat"}},
  {{"subject": "Chương trình giáo dục phổ thông", "subject_type": "ChuongTrinhHocLieu", "predicate": "phai_dap_ung", "object": "tổ chức thực hiện linh hoạt", "object_type": "YeuCauTinhChat"}},
  {{"subject": "Chương trình giáo dục phổ thông", "subject_type": "ChuongTrinhHocLieu", "predicate": "phai_dap_ung", "object": "công bố công khai", "object_type": "YeuCauTinhChat"}}
]


### NHIỆM VỤ CỦA BẠN:
####
Tài liệu: {document}
####
Ontology: {ontology}
####

KG Triplets:"""

    return extract_json_array(call_model(generator, prompt, max_tokens))


def build_node_id(khoan, dieu) -> str:
    parts = []
    if pd.notna(khoan) and str(khoan).strip() not in ("", "nan"):
        val_khoan = int(khoan) if isinstance(khoan, (int, float)) else str(khoan).strip()
        parts.append(f"khoản{val_khoan}")
    if pd.notna(dieu) and str(dieu).strip() not in ("", "nan"):
        val_dieu = int(dieu) if isinstance(dieu, (int, float)) else str(dieu).strip()
        parts.append(f"điều{val_dieu}")
    parts.append(LAW_ID)
    return "-".join(parts)


def extract_first_subject_from_kg(kg_json_str: str, fallback_document: str) -> str:
    """
    Trích xuất subject đầu tiên từ KG JSON do LM sinh ra.
    Nếu parse JSON thất bại, dùng cụm từ đầu tiên của document làm fallback.
    """
    try:
        match = re.search(r'\[[\s\S]*\]', kg_json_str)
        if match:
            triplets = json.loads(match.group(0))
            if triplets and isinstance(triplets, list):
                first = triplets[0]
                subject = first.get("subject", "").strip()
                if subject:
                    return subject
    except (json.JSONDecodeError, KeyError, IndexError):
        pass

    fallback = re.split(r'[,;.]', fallback_document.strip())[0].strip()
    return fallback[:80] if fallback else "Chủ thể không xác định"


def extract_first_subject_type_from_kg(kg_json_str: str, fallback: str = "UNKNOWN") -> str:
    """
    Trích xuất subject_type của subject đầu tiên từ KG JSON do LM sinh ra.
    Nếu parse JSON thất bại, trả về fallback mặc định là "UNKNOWN".
    """
    try:
        match = re.search(r'\[[\s\S]*\]', kg_json_str)
        if match:
            triplets = json.loads(match.group(0))
            if triplets and isinstance(triplets, list):
                first = triplets[0]
                subject_type = first.get("subject_type", "").strip()
                if subject_type:
                    return subject_type
    except (json.JSONDecodeError, KeyError, IndexError):
        pass

    return fallback


def create_fixed_triplets(khoan, dieu, noi_dung: str, first_subject: str, first_subject_type: str) -> list:
    """
    Tạo 3 triplet cố định để gắn kết toàn bộ bộ luật:

    1. (luật 43/2019/QH14, bao_gom, khoản{k}-điều{d}-43/2019/QH14)
    2. (khoản{k}-điều{d}-43/2019/QH14, co_noi_dung, {nội dung})
    3. (khoản{k}-điều{d}-43/2019/QH14, quy_dinh, {subject đầu tiên})
    """
    node_id = build_node_id(khoan, dieu)

    # Xác định node_type dựa trên có khoản hay không
    if pd.notna(khoan) and str(khoan).strip() not in ("", "nan"):
        node_type = "CauTrucKhoan"
    else:
        node_type = "CauTrucDieu"

    fixed = [
        {
            "subject": f"luật {LAW_ID}",
            "subject_type": "VanBanLuat",
            "predicate": "bao_gom",
            "object": node_id,
            "object_type": node_type
        },
        {
            "subject": node_id,
            "subject_type": node_type,
            "predicate": "co_noi_dung",
            "object": str(noi_dung).strip(),
            "object_type": "UNKNOWN"
        },
        {
            "subject": node_id,
            "subject_type": node_type,
            "predicate": "quy_dinh",
            "object": first_subject,
            "object_type": first_subject_type
        }
    ]
    return fixed


def KG_triplet_pipeline(generator, document: str,
                        khoan=None, dieu=None, noi_dung: str = "") -> dict:

    # --- Tạo KG Triplets bằng LM ---
    triplets_text = create_KG(generator, document, ontology)

    # --- Parse LM triplets thành list dict ---
    lm_triplets = []
    try:
        match = re.search(r'\[[\s\S]*\]', triplets_text)
        if match:
            lm_triplets = json.loads(match.group(0))
    except (json.JSONDecodeError, AttributeError):
        print("[WARN] Không parse được JSON từ LM output, lưu dạng raw string.")

    # --- Trích xuất subject đầu tiên từ KG của LM ---
    first_subject = extract_first_subject_from_kg(triplets_text, document)
    first_subject_type = extract_first_subject_type_from_kg(triplets_text)

    # --- Tạo 3 triplet cố định ---
    fixed_triplets = create_fixed_triplets(khoan, dieu, noi_dung, first_subject, first_subject_type)

    # --- Gộp tất cả triplets ---
    all_triplets = fixed_triplets + (lm_triplets if isinstance(lm_triplets, list) else [])

    return {
        "triplets_lm_raw": triplets_text,
        "triplets_fixed": fixed_triplets,
        "triplets_all": all_triplets
    }

def export_excel(df: pd.DataFrame, all_results: list):
    triplet_col = []
    json_col    = []

    for res in all_results:
        lines = []
        for t in res.get("triplets_all", []):
            s = t.get("subject", "")
            p = t.get("predicate", "")
            o = t.get("object", "")
            lines.append(f"({s}, {p}, {o})")
        triplet_col.append("\n".join(lines))
        json_col.append(res.get("triplets_all", ""))

    df_out = df.copy()
    df_out["Triplet"]   = triplet_col
    df_out["Json gốc"] = json_col

    df_out.to_excel(OUTPUT_EXCEL, index=False, engine="openpyxl")

    print(f"✅ Đã xuất {len(df_out)} dòng ra: {OUTPUT_EXCEL}")

def main():
    generator = load_model(MODEL_NAME)

    df = pd.read_excel("Data-Sub.xlsx")

    all_results = []

    for idx, row in df.iterrows():
        doc_id = row.get(COL_DOC_ID, idx + 1)
        document = str(row[COL_DOCUMENT]).strip()
        khoan = row.get(COL_KHOAN)
        dieu = row.get(COL_DIEU)
        noi_dung = str(row[COL_NOI_DUNG]).strip() if COL_NOI_DUNG in row and pd.notna(row[COL_NOI_DUNG]) else document

        print("\n" + "🔷" * 35)
        print(f"  📄 DOC {doc_id}  |  Khoản: {khoan}  |  Điều: {dieu}")
        print("🔷" * 35)
        print(document)

        result = KG_triplet_pipeline(
            generator,
            document=document,
            khoan=khoan,
            dieu=dieu,
            noi_dung=noi_dung
        )

        all_results.append({
            "doc_id": doc_id,
            "document": document,
            "khoan": khoan,
            "dieu": dieu,
            **result
        })

        if (idx + 1) % SAVE_EVERY_N_ROWS == 0:
            print(f"\n[HỆ THỐNG] Đang lưu tạm thời kết quả tại dòng {idx + 1}...")
            df_subset = df.iloc[:len(all_results)]
            export_excel(df_subset, all_results)

    print("\n✅ Hoàn thành toàn bộ pipeline!")
    export_excel(df, all_results)


if __name__ == "__main__":
    main()